"""
=============================================================================
  ACERVO TECH DIGITAL — Back-end completo com MySQL
  Tecnologia: FastAPI + MySQL (pymysql)
  Funcionalidades: CRUD, gráficos, busca, QR Code, alertas, lixeira,
                   histórico, relatórios PDF/Excel, lembretes por e-mail,
                   RESERVAS DE LIVROS (com fila de espera)
=============================================================================
"""

import hashlib
import pymysql
from datetime import date, datetime, timedelta
from fastapi import FastAPI, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pathlib import Path
BASE_DIR = Path(__file__).resolve().parent.parent  # sobe de back-end para acervotech
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, field_validator
from typing import Optional
import qrcode
from io import BytesIO
import os
import smtplib
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger
import logging

# ===========================================================================
#  BIBLIOTECAS PARA RELATÓRIOS
# ===========================================================================
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile

# ---------------------------------------------------------------------------
# CONFIGURAÇÕES DO BANCO DE DADOS (porta 3307)
# ---------------------------------------------------------------------------
DB_CONFIG = {
    "host": "localhost",
    "port": 3306,
    "user": "root",
    "password": "12345678",
    "database": "acervotech",
    "charset": "utf8mb4",
    "cursorclass": pymysql.cursors.DictCursor
}

# ---------------------------------------------------------------------------
# CONFIGURAÇÕES DE E-MAIL (ALTERE PARA SEUS DADOS)
# ---------------------------------------------------------------------------
EMAIL_CONFIG = {
    "smtp_server": "smtp.gmail.com",
    "smtp_port": 587,
    "email": "seuemail@gmail.com",       # ← SEU E-MAIL
    "senha": "sua_senha_de_app",          # ← SENHA DE APP
    "assunto_prefixo": "[AcervoTech] "
}

# ---------------------------------------------------------------------------
# Instância principal do FastAPI e CORS
# ---------------------------------------------------------------------------
app = FastAPI(
    title="Acervo Tech Digital",
    description="API de gerenciamento de biblioteca digital com MySQL",
    version="2.1.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ===========================================================================
#  UTILITÁRIOS
# ===========================================================================

def hash_senha(senha: str) -> str:
    return hashlib.sha256(senha.encode()).hexdigest()

def obter_conexao():
    conn = pymysql.connect(**DB_CONFIG)
    return conn

# ===========================================================================
#  INICIALIZAÇÃO DO BANCO DE DADOS
# ===========================================================================

def inicializar_banco():
    conn = obter_conexao()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS admins (
            id     INT PRIMARY KEY AUTO_INCREMENT,
            nome   VARCHAR(100) NOT NULL,
            email  VARCHAR(100) UNIQUE NOT NULL,
            senha  VARCHAR(64) NOT NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS livros (
            id              INT PRIMARY KEY AUTO_INCREMENT,
            titulo          VARCHAR(200) NOT NULL,
            autor           VARCHAR(100) NOT NULL,
            isbn            VARCHAR(20),
            genero          VARCHAR(50),
            exemplares      INT DEFAULT 1,
            disponiveis     INT DEFAULT 1,
            ativo           BOOLEAN DEFAULT TRUE,
            data_inativacao DATE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS usuarios_biblio (
            id    INT PRIMARY KEY AUTO_INCREMENT,
            nome  VARCHAR(100) NOT NULL,
            email VARCHAR(100) UNIQUE NOT NULL,
            tipo  VARCHAR(20) NOT NULL DEFAULT 'Aluno',
            ativo BOOLEAN DEFAULT TRUE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS emprestimos (
            id                      INT PRIMARY KEY AUTO_INCREMENT,
            id_livro                INT NOT NULL,
            id_usuario              INT NOT NULL,
            data_emprestimo         DATE NOT NULL,
            data_devolucao_prevista DATE NOT NULL,
            data_devolucao_real     DATE,
            status                  VARCHAR(20) NOT NULL DEFAULT 'ativo',
            FOREIGN KEY (id_livro)   REFERENCES livros(id) ON DELETE RESTRICT,
            FOREIGN KEY (id_usuario) REFERENCES usuarios_biblio(id) ON DELETE RESTRICT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)

    # ============================================================
    # NOVA TABELA: RESERVAS
    # ============================================================
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS reservas (
            id              INT PRIMARY KEY AUTO_INCREMENT,
            id_livro        INT NOT NULL,
            id_usuario      INT NOT NULL,
            data_reserva    DATE NOT NULL,
            status          VARCHAR(20) NOT NULL DEFAULT 'ativa',  -- 'ativa', 'cancelada', 'concluida'
            FOREIGN KEY (id_livro)   REFERENCES livros(id) ON DELETE RESTRICT,
            FOREIGN KEY (id_usuario) REFERENCES usuarios_biblio(id) ON DELETE RESTRICT
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
    """)

    # Verifica se já existe um admin padrão
    cursor.execute("SELECT COUNT(*) AS total FROM admins")
    total = cursor.fetchone()["total"]
    if total == 0:
        cursor.execute(
            "INSERT INTO admins (nome, email, senha) VALUES (%s, %s, %s)",
            ("Admin Master", "admin@teste.com", hash_senha("123456"))
        )

    conn.commit()
    conn.close()
    print("✅ Tabelas verificadas/criadas com sucesso no MySQL!")

inicializar_banco()

# ===========================================================================
#  MODELOS PYDANTIC
# ===========================================================================

class ModeloCadastroAdmin(BaseModel):
    nome:  str
    email: str
    senha: str

class ModeloLogin(BaseModel):
    email: str
    senha: str

class ModeloLivro(BaseModel):
    titulo:     str
    autor:      str
    isbn:       Optional[str] = None
    genero:     Optional[str] = None
    exemplares: int = 1

    @field_validator("exemplares")
    @classmethod
    def exemplares_positivos(cls, v: int) -> int:
        if v < 1:
            raise ValueError("A quantidade de exemplares deve ser pelo menos 1.")
        return v

class ModeloUsuarioBiblio(BaseModel):
    nome:  str
    email: str
    tipo:  str = "Aluno"
    ativo: bool = True

class ModeloEmprestimo(BaseModel):
    id_livro:               int
    id_usuario:             int
    data_devolucao_prevista: str

    @field_validator("data_devolucao_prevista")
    @classmethod
    def validar_data_futura(cls, v: str) -> str:
        try:
            data = datetime.strptime(v, "%Y-%m-%d").date()
        except ValueError:
            raise ValueError("Data deve estar no formato YYYY-MM-DD.")
        if data < date.today():
            raise ValueError("A data de devolução prevista não pode ser anterior a hoje.")
        return v

class ModeloReserva(BaseModel):
    id_livro:   int
    id_usuario: int

# ===========================================================================
#  ROTAS — AUTENTICAÇÃO
# ===========================================================================

@app.post("/api/auth/register", status_code=status.HTTP_201_CREATED)
def cadastrar_admin(dados: ModeloCadastroAdmin):
    conn = obter_conexao()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO admins (nome, email, senha) VALUES (%s, %s, %s)",
            (dados.nome, dados.email, hash_senha(dados.senha))
        )
        conn.commit()
        return {"message": "Administrador cadastrado com sucesso!"}
    except pymysql.IntegrityError:
        raise HTTPException(status_code=400, detail="Este e-mail já está cadastrado.")
    finally:
        conn.close()

@app.post("/api/auth/login")
def login_admin(dados: ModeloLogin):
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, nome, email FROM admins WHERE email = %s AND senha = %s",
        (dados.email, hash_senha(dados.senha))
    )
    admin = cursor.fetchone()
    conn.close()
    if admin:
        return {
            "message": "Login efetuado com sucesso!",
            "usuario": {
                "id":    admin["id"],
                "nome":  admin["nome"],
                "email": admin["email"]
            }
        }
    raise HTTPException(status_code=401, detail="E-mail ou senha incorretos.")

# ===========================================================================
#  ROTAS — LIVROS (com busca)
# ===========================================================================

@app.post("/api/livros", status_code=status.HTTP_201_CREATED)
def cadastrar_livro(livro: ModeloLivro):
    conn = obter_conexao()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """INSERT INTO livros (titulo, autor, isbn, genero, exemplares, disponiveis)
               VALUES (%s, %s, %s, %s, %s, %s)""",
            (livro.titulo, livro.autor, livro.isbn,
             livro.genero, livro.exemplares, livro.exemplares)
        )
        conn.commit()
        return {"message": "Livro cadastrado com sucesso!", "id": cursor.lastrowid}
    except Exception as erro:
        raise HTTPException(status_code=500, detail=f"Erro ao salvar livro: {str(erro)}")
    finally:
        conn.close()

@app.get("/api/livros", summary="Listar livros ativos (com busca)")
def listar_livros(q: Optional[str] = None):
    conn = obter_conexao()
    cursor = conn.cursor()
    if q:
        like = f"%{q}%"
        cursor.execute(
            """SELECT id, titulo, autor, isbn, genero, exemplares, disponiveis
               FROM livros
               WHERE ativo = 1 AND (titulo LIKE %s OR autor LIKE %s OR isbn LIKE %s)
               ORDER BY titulo""",
            (like, like, like)
        )
    else:
        cursor.execute(
            "SELECT id, titulo, autor, isbn, genero, exemplares, disponiveis "
            "FROM livros WHERE ativo = 1 ORDER BY titulo"
        )
    linhas = cursor.fetchall()
    conn.close()
    return linhas

@app.get("/api/livros/lixeira")
def listar_livros_lixeira():
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, titulo, autor, isbn, genero, exemplares, disponiveis, data_inativacao "
        "FROM livros WHERE ativo = 0 ORDER BY data_inativacao DESC"
    )
    linhas = cursor.fetchall()
    conn.close()
    return linhas

@app.get("/api/livros/{livro_id}")
def buscar_livro(livro_id: int):
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM livros WHERE id = %s", (livro_id,))
    livro = cursor.fetchone()
    conn.close()
    if not livro:
        raise HTTPException(status_code=404, detail="Livro não encontrado.")
    return livro

@app.put("/api/livros/{livro_id}")
def atualizar_livro(livro_id: int, livro: ModeloLivro):
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT exemplares, disponiveis FROM livros WHERE id = %s", (livro_id,))
    livro_atual = cursor.fetchone()
    if not livro_atual:
        conn.close()
        raise HTTPException(status_code=404, detail="Livro não encontrado.")
    variacao = livro.exemplares - livro_atual["exemplares"]
    novos_disponiveis = livro_atual["disponiveis"] + variacao
    if novos_disponiveis < 0:
        conn.close()
        raise HTTPException(status_code=400, detail="Não é possível reduzir abaixo do emprestado.")
    try:
        cursor.execute(
            """UPDATE livros
               SET titulo=%s, autor=%s, isbn=%s, genero=%s, exemplares=%s, disponiveis=%s
               WHERE id=%s""",
            (livro.titulo, livro.autor, livro.isbn, livro.genero,
             livro.exemplares, novos_disponiveis, livro_id)
        )
        conn.commit()
        return {"message": "Livro atualizado com sucesso!"}
    except Exception as erro:
        raise HTTPException(status_code=500, detail=f"Erro ao atualizar: {str(erro)}")
    finally:
        conn.close()

@app.delete("/api/livros/{livro_id}")
def mover_para_lixeira(livro_id: int):
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT ativo FROM livros WHERE id = %s", (livro_id,))
    livro = cursor.fetchone()
    if not livro or not livro["ativo"]:
        conn.close()
        raise HTTPException(status_code=400, detail="Livro não encontrado ou já na lixeira.")
    cursor.execute("SELECT id FROM emprestimos WHERE id_livro = %s AND status = 'ativo'", (livro_id,))
    if cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="Existem empréstimos ativos para este livro.")
    data_atual = date.today().isoformat()
    cursor.execute(
        "UPDATE livros SET ativo = 0, data_inativacao = %s WHERE id = %s",
        (data_atual, livro_id)
    )
    conn.commit()
    conn.close()
    return {"message": "Livro movido para a lixeira."}

@app.put("/api/livros/{livro_id}/restaurar")
def restaurar_livro(livro_id: int):
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT ativo FROM livros WHERE id = %s", (livro_id,))
    livro = cursor.fetchone()
    if not livro or livro["ativo"]:
        conn.close()
        raise HTTPException(status_code=400, detail="Livro não encontrado ou já ativo.")
    cursor.execute("UPDATE livros SET ativo = 1, data_inativacao = NULL WHERE id = %s", (livro_id,))
    conn.commit()
    conn.close()
    return {"message": "Livro restaurado com sucesso!"}

@app.delete("/api/livros/limpeza-definitiva")
def limpeza_definitiva():
    conn = obter_conexao()
    cursor = conn.cursor()
    cinco_anos_atras = (date.today() - timedelta(days=1825)).isoformat()
    cursor.execute(
        "DELETE FROM livros WHERE ativo = 0 AND data_inativacao <= %s",
        (cinco_anos_atras,)
    )
    conn.commit()
    conn.close()
    return {"message": "Limpeza concluída!"}

# ===========================================================================
#  ROTAS — USUÁRIOS
# ===========================================================================

@app.post("/api/usuarios", status_code=status.HTTP_201_CREATED)
def cadastrar_usuario_biblio(usuario: ModeloUsuarioBiblio):
    conn = obter_conexao()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO usuarios_biblio (nome, email, tipo, ativo) VALUES (%s, %s, %s, %s)",
            (usuario.nome, usuario.email, usuario.tipo, int(usuario.ativo))
        )
        conn.commit()
        return {"message": "Usuário cadastrado com sucesso!", "id": cursor.lastrowid}
    except pymysql.IntegrityError:
        raise HTTPException(status_code=400, detail="Este e-mail já está cadastrado.")
    finally:
        conn.close()

@app.get("/api/usuarios")
def listar_usuarios():
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT id, nome, email, tipo, ativo FROM usuarios_biblio ORDER BY nome")
    linhas = cursor.fetchall()
    conn.close()
    for u in linhas:
        u["ativo"] = bool(u["ativo"])
    return linhas

@app.put("/api/usuarios/{usuario_id}")
def atualizar_usuario_biblio(usuario_id: int, usuario: ModeloUsuarioBiblio):
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM usuarios_biblio WHERE id = %s", (usuario_id,))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")
    cursor.execute(
        "UPDATE usuarios_biblio SET nome=%s, email=%s, tipo=%s, ativo=%s WHERE id=%s",
        (usuario.nome, usuario.email, usuario.tipo, int(usuario.ativo), usuario_id)
    )
    conn.commit()
    conn.close()
    return {"message": "Usuário atualizado com sucesso!"}

@app.delete("/api/usuarios/{usuario_id}")
def excluir_usuario_biblio(usuario_id: int):
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM usuarios_biblio WHERE id = %s", (usuario_id,))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")
    cursor.execute("SELECT id FROM emprestimos WHERE id_usuario = %s AND status = 'ativo'", (usuario_id,))
    if cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="Usuário possui empréstimos ativos.")
    cursor.execute("DELETE FROM usuarios_biblio WHERE id = %s", (usuario_id,))
    conn.commit()
    conn.close()
    return {"message": "Usuário excluído com sucesso!"}

# ===========================================================================
#  ROTA — HISTÓRICO DO USUÁRIO
# ===========================================================================

@app.get("/api/usuarios/{usuario_id}/historico", summary="Histórico de empréstimos do usuário")
def historico_usuario(usuario_id: int):
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT id, nome FROM usuarios_biblio WHERE id = %s", (usuario_id,))
    usuario = cursor.fetchone()
    if not usuario:
        conn.close()
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")
    cursor.execute("""
        SELECT e.id,
               l.titulo AS titulo_livro,
               e.data_emprestimo,
               e.data_devolucao_prevista,
               e.data_devolucao_real,
               e.status,
               DATEDIFF(e.data_devolucao_real, e.data_devolucao_prevista) AS dias_atraso
          FROM emprestimos e
          JOIN livros l ON l.id = e.id_livro
         WHERE e.id_usuario = %s
         ORDER BY e.data_emprestimo DESC
    """, (usuario_id,))
    emprestimos = cursor.fetchall()
    conn.close()
    return {
        "usuario": usuario,
        "emprestimos": emprestimos,
        "total": len(emprestimos)
    }

# ===========================================================================
#  ROTAS — EMPRÉSTIMOS (com busca)
# ===========================================================================

@app.post("/api/emprestimos", status_code=status.HTTP_201_CREATED)
def registrar_emprestimo(dados: ModeloEmprestimo):
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT disponiveis FROM livros WHERE id = %s AND ativo = 1", (dados.id_livro,))
    livro = cursor.fetchone()
    if not livro or livro["disponiveis"] <= 0:
        conn.close()
        raise HTTPException(status_code=400, detail="Livro indisponível.")
    cursor.execute("SELECT ativo FROM usuarios_biblio WHERE id = %s", (dados.id_usuario,))
    usuario = cursor.fetchone()
    if not usuario or not usuario["ativo"]:
        conn.close()
        raise HTTPException(status_code=400, detail="Usuário inativo ou não encontrado.")
    hoje = date.today().isoformat()
    cursor.execute(
        """INSERT INTO emprestimos
               (id_livro, id_usuario, data_emprestimo, data_devolucao_prevista, status)
           VALUES (%s, %s, %s, %s, 'ativo')""",
        (dados.id_livro, dados.id_usuario, hoje, dados.data_devolucao_prevista)
    )
    cursor.execute("UPDATE livros SET disponiveis = disponiveis - 1 WHERE id = %s", (dados.id_livro,))
    conn.commit()
    novo_id = cursor.lastrowid
    conn.close()
    return {"message": "Empréstimo registrado com sucesso!", "id": novo_id}

@app.get("/api/emprestimos", summary="Listar empréstimos (com busca)")
def listar_emprestimos(q: Optional[str] = None, status_filtro: Optional[str] = None):
    conn = obter_conexao()
    cursor = conn.cursor()
    sql = """
        SELECT e.id,
               e.id_livro,
               l.titulo          AS titulo_livro,
               e.id_usuario,
               u.nome            AS nome_usuario,
               e.data_emprestimo,
               e.data_devolucao_prevista,
               e.data_devolucao_real,
               e.status
          FROM emprestimos e
          JOIN livros            l ON l.id = e.id_livro
          JOIN usuarios_biblio   u ON u.id = e.id_usuario
         WHERE 1=1
    """
    params = []
    if q:
        sql += " AND (l.titulo LIKE %s OR u.nome LIKE %s)"
        like = f"%{q}%"
        params.extend([like, like])
    if status_filtro:
        sql += " AND e.status = %s"
        params.append(status_filtro)
    sql += " ORDER BY e.id DESC"
    cursor.execute(sql, params)
    linhas = cursor.fetchall()
    conn.close()
    return linhas

@app.post("/api/emprestimos/{emprestimo_id}/devolver")
def registrar_devolucao(emprestimo_id: int):
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM emprestimos WHERE id = %s", (emprestimo_id,))
    emprestimo = cursor.fetchone()
    if not emprestimo or emprestimo["status"] == "finalizado":
        conn.close()
        raise HTTPException(status_code=400, detail="Empréstimo não encontrado ou já finalizado.")
    hoje = date.today()
    previsao = datetime.strptime(emprestimo["data_devolucao_prevista"], "%Y-%m-%d").date()
    dias_atraso = max(0, (hoje - previsao).days)
    multa = round(dias_atraso * 2.00, 2)
    
    # Iniciar transação
    try:
        # 1. Atualizar empréstimo
        cursor.execute(
            "UPDATE emprestimos SET status = 'finalizado', data_devolucao_real = %s WHERE id = %s",
            (hoje.isoformat(), emprestimo_id)
        )
        # 2. Incrementar disponíveis
        cursor.execute(
            "UPDATE livros SET disponiveis = LEAST(disponiveis + 1, exemplares) WHERE id = %s",
            (emprestimo["id_livro"],)
        )
        # 3. Verificar se há reservas para este livro (a mais antiga)
        cursor.execute("""
            SELECT r.id, r.id_usuario, u.nome, u.email, l.titulo
              FROM reservas r
              JOIN usuarios_biblio u ON u.id = r.id_usuario
              JOIN livros l ON l.id = r.id_livro
             WHERE r.id_livro = %s AND r.status = 'ativa'
             ORDER BY r.data_reserva ASC
             LIMIT 1
        """, (emprestimo["id_livro"],))
        reserva = cursor.fetchone()
        
        if reserva:
            # Enviar e-mail para o usuário da reserva
            assunto = f"Livro disponível: {reserva['titulo']}"
            corpo = f"""
Olá {reserva['nome']},

O livro "{reserva['titulo']}" que você reservou foi devolvido e agora está disponível.

Por favor, dirija-se à biblioteca para efetuar o empréstimo o mais breve possível.

Atenciosamente,
Equipe AcervoTech
            """
            enviar_email(reserva['email'], assunto, corpo)
            
            # Opcional: marcar reserva como concluída (ou manter ativa até o empréstimo)
            # Vamos manter ativa para controle, mas podemos marcar como 'concluida' quando o empréstimo for efetuado.
            # Não faremos automaticamente para permitir que o bibliotecário decida.
        
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

    return {
        "message": "Devolução registrada com sucesso!",
        "dias_atraso": dias_atraso,
        "multa": multa,
        "reserva_notificada": bool(reserva)
    }

# ===========================================================================
#  ROTAS — RESERVAS (NOVO)
# ===========================================================================

@app.post("/api/reservas", status_code=status.HTTP_201_CREATED, summary="Criar uma reserva")
def criar_reserva(reserva: ModeloReserva):
    """
    Cria uma reserva para um livro. Verifica se o livro está indisponível (disponiveis = 0)
    e se o usuário não tem reserva ativa para o mesmo livro.
    """
    conn = obter_conexao()
    cursor = conn.cursor()
    
    # Verificar se o livro existe e está ativo
    cursor.execute("SELECT id, disponiveis, titulo FROM livros WHERE id = %s AND ativo = 1", (reserva.id_livro,))
    livro = cursor.fetchone()
    if not livro:
        conn.close()
        raise HTTPException(status_code=404, detail="Livro não encontrado ou inativo.")
    
    # Verificar se o usuário existe e está ativo
    cursor.execute("SELECT id, nome, email FROM usuarios_biblio WHERE id = %s AND ativo = 1", (reserva.id_usuario,))
    usuario = cursor.fetchone()
    if not usuario:
        conn.close()
        raise HTTPException(status_code=404, detail="Usuário não encontrado ou inativo.")
    
    # Verificar se o livro está disponível (se disponiveis > 0, não precisa reservar)
    if livro["disponiveis"] > 0:
        conn.close()
        raise HTTPException(status_code=400, detail="Livro está disponível. Não é necessário reservar.")
    
    # Verificar se o usuário já tem reserva ativa para este livro
    cursor.execute(
        "SELECT id FROM reservas WHERE id_livro = %s AND id_usuario = %s AND status = 'ativa'",
        (reserva.id_livro, reserva.id_usuario)
    )
    if cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="Usuário já possui reserva ativa para este livro.")
    
    # Criar a reserva
    hoje = date.today().isoformat()
    cursor.execute(
        "INSERT INTO reservas (id_livro, id_usuario, data_reserva, status) VALUES (%s, %s, %s, 'ativa')",
        (reserva.id_livro, reserva.id_usuario, hoje)
    )
    conn.commit()
    novo_id = cursor.lastrowid
    conn.close()
    
    # Opcional: enviar e-mail de confirmação
    enviar_email(
        usuario["email"],
        f"Reserva confirmada: {livro['titulo']}",
        f"Olá {usuario['nome']},\n\nSua reserva para o livro '{livro['titulo']}' foi registrada com sucesso.\n\nAguardamos a devolução para que você possa retirá-lo.\n\nAtenciosamente,\nEquipe AcervoTech"
    )
    
    return {"message": "Reserva criada com sucesso!", "id": novo_id}


@app.get("/api/reservas", summary="Listar reservas (com filtros)")
def listar_reservas(status: Optional[str] = None, id_livro: Optional[int] = None, id_usuario: Optional[int] = None):
    """
    Lista todas as reservas com opção de filtrar por status, livro ou usuário.
    """
    conn = obter_conexao()
    cursor = conn.cursor()
    sql = """
        SELECT r.id,
               r.id_livro,
               l.titulo AS titulo_livro,
               r.id_usuario,
               u.nome   AS nome_usuario,
               r.data_reserva,
               r.status
          FROM reservas r
          JOIN livros l ON l.id = r.id_livro
          JOIN usuarios_biblio u ON u.id = r.id_usuario
         WHERE 1=1
    """
    params = []
    if status:
        sql += " AND r.status = %s"
        params.append(status)
    if id_livro:
        sql += " AND r.id_livro = %s"
        params.append(id_livro)
    if id_usuario:
        sql += " AND r.id_usuario = %s"
        params.append(id_usuario)
    sql += " ORDER BY r.data_reserva ASC"
    
    cursor.execute(sql, params)
    reservas = cursor.fetchall()
    conn.close()
    return reservas


@app.delete("/api/reservas/{reserva_id}", summary="Cancelar uma reserva")
def cancelar_reserva(reserva_id: int):
    """
    Cancela uma reserva, alterando o status para 'cancelada'.
    """
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT id, status FROM reservas WHERE id = %s", (reserva_id,))
    reserva = cursor.fetchone()
    if not reserva:
        conn.close()
        raise HTTPException(status_code=404, detail="Reserva não encontrada.")
    if reserva["status"] != "ativa":
        conn.close()
        raise HTTPException(status_code=400, detail="Esta reserva já foi cancelada ou concluída.")
    
    cursor.execute("UPDATE reservas SET status = 'cancelada' WHERE id = %s", (reserva_id,))
    conn.commit()
    conn.close()
    return {"message": "Reserva cancelada com sucesso."}


@app.get("/api/reservas/por-usuario/{usuario_id}", summary="Listar reservas de um usuário")
def reservas_por_usuario(usuario_id: int):
    """
    Retorna todas as reservas ativas e concluídas de um usuário específico.
    """
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT r.id,
               r.id_livro,
               l.titulo AS titulo_livro,
               r.data_reserva,
               r.status
          FROM reservas r
          JOIN livros l ON l.id = r.id_livro
         WHERE r.id_usuario = %s
         ORDER BY r.data_reserva DESC
    """, (usuario_id,))
    reservas = cursor.fetchall()
    conn.close()
    return reservas

# ===========================================================================
#  ROTAS — DASHBOARD E GRÁFICOS
# ===========================================================================

@app.get("/api/dashboard")
def obter_estatisticas():
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) AS total FROM livros WHERE ativo = 1")
    total_livros = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM usuarios_biblio WHERE ativo = 1")
    total_usuarios = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM emprestimos WHERE status = 'ativo'")
    total_emprestimos = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM livros WHERE ativo = 1 AND disponiveis = 0")
    total_indisponiveis = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM reservas WHERE status = 'ativa'")
    total_reservas = cursor.fetchone()["total"]
    conn.close()
    return {
        "total_livros": total_livros,
        "total_usuarios": total_usuarios,
        "total_emprestimos": total_emprestimos,
        "total_indisponiveis": total_indisponiveis,
        "total_reservas": total_reservas
    }

@app.get("/api/estatisticas/graficos")
def obter_dados_graficos():
    conn = obter_conexao()
    cursor = conn.cursor()
    # Empréstimos por mês (últimos 6 meses)
    meses = []
    hoje = date.today()
    for i in range(5, -1, -1):
        mes = hoje.month - i
        ano = hoje.year
        if mes <= 0:
            mes += 12
            ano -= 1
        meses.append(f"{ano}-{mes:02d}")

    dados_mensais = []
    for mes in meses:
        cursor.execute(
            "SELECT COUNT(*) AS total FROM emprestimos WHERE DATE_FORMAT(data_emprestimo, '%Y-%m') = %s",
            (mes,)
        )
        dados_mensais.append(cursor.fetchone()["total"])

    nomes_meses = []
    for mes in meses:
        ano, mes_num = mes.split("-")
        nome_mes = datetime.strptime(mes_num, "%m").strftime("%b")
        nomes_meses.append(f"{nome_mes}/{ano}")

    # Livros por gênero
    cursor.execute("""
        SELECT genero, COUNT(*) AS total
        FROM livros
        WHERE ativo = 1 AND genero IS NOT NULL AND genero != ''
        GROUP BY genero
        ORDER BY total DESC
        LIMIT 5
    """)
    generos = cursor.fetchall()
    conn.close()
    return {
        "emprestimos_por_mes": {
            "labels": nomes_meses,
            "valores": dados_mensais
        },
        "livros_por_genero": {
            "labels": [g["genero"] for g in generos],
            "valores": [g["total"] for g in generos]
        }
    }

# ===========================================================================
#  ROTA — ALERTAS DE DEVOLUÇÃO VENCIDA
# ===========================================================================

@app.get("/api/alertas/vencidos", summary="Alertas de devoluções vencidas")
def alertas_devolucao_vencida():
    conn = obter_conexao()
    cursor = conn.cursor()
    hoje = date.today().isoformat()
    cursor.execute("""
        SELECT e.id,
               l.titulo AS titulo_livro,
               u.nome   AS nome_usuario,
               e.data_devolucao_prevista,
               DATEDIFF(%s, e.data_devolucao_prevista) AS dias_atraso
          FROM emprestimos e
          JOIN livros l ON l.id = e.id_livro
          JOIN usuarios_biblio u ON u.id = e.id_usuario
         WHERE e.status = 'ativo'
           AND e.data_devolucao_prevista < %s
         ORDER BY e.data_devolucao_prevista ASC
    """, (hoje, hoje))
    vencidos = cursor.fetchall()
    conn.close()
    return {
        "vencidos": vencidos,
        "total": len(vencidos)
    }

# ===========================================================================
#  ROTAS — RELATÓRIOS (PDF e EXCEL)
# ===========================================================================

@app.get("/api/relatorios/livros/pdf", summary="Exportar livros para PDF")
def relatorio_livros_pdf():
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, titulo, autor, isbn, genero, exemplares, disponiveis "
        "FROM livros WHERE ativo = 1 ORDER BY titulo"
    )
    livros = cursor.fetchall()
    conn.close()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        pdf_path = tmp.name

    doc = SimpleDocTemplate(pdf_path, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    styles = getSampleStyleSheet()
    titulo = Paragraph("📚 Relatório de Livros - AcervoTech", styles['Title'])
    elements.append(titulo)
    elements.append(Spacer(1, 0.5*cm))
    subtitulo = Paragraph(f"Gerado em: {date.today().strftime('%d/%m/%Y')}", styles['Normal'])
    elements.append(subtitulo)
    elements.append(Spacer(1, 1*cm))

    dados = [["ID", "Título", "Autor", "ISBN", "Gênero", "Exemplares", "Disponíveis"]]
    for livro in livros:
        dados.append([
            str(livro["id"]),
            livro["titulo"],
            livro["autor"],
            livro["isbn"] or "-",
            livro["genero"] or "-",
            str(livro["exemplares"]),
            str(livro["disponiveis"])
        ])

    tabela = Table(dados, colWidths=[1.5*cm, 5*cm, 4*cm, 3*cm, 3*cm, 2.5*cm, 2.5*cm])
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a3f8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
    ]))
    elements.append(tabela)
    elements.append(Spacer(1, 1*cm))
    total = Paragraph(f"Total de livros: {len(livros)}", styles['Normal'])
    elements.append(total)
    doc.build(elements)

    return FileResponse(pdf_path, media_type='application/pdf', filename=f"relatorio_livros_{date.today().isoformat()}.pdf")


@app.get("/api/relatorios/livros/excel", summary="Exportar livros para Excel")
def relatorio_livros_excel():
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, titulo, autor, isbn, genero, exemplares, disponiveis "
        "FROM livros WHERE ativo = 1 ORDER BY titulo"
    )
    livros = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Livros"
    headers = ["ID", "Título", "Autor", "ISBN", "Gênero", "Exemplares", "Disponíveis"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4a3f8a", end_color="4a3f8a", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, livro in enumerate(livros, 2):
        ws.cell(row=row, column=1, value=livro["id"])
        ws.cell(row=row, column=2, value=livro["titulo"])
        ws.cell(row=row, column=3, value=livro["autor"])
        ws.cell(row=row, column=4, value=livro["isbn"] or "-")
        ws.cell(row=row, column=5, value=livro["genero"] or "-")
        ws.cell(row=row, column=6, value=livro["exemplares"])
        ws.cell(row=row, column=7, value=livro["disponiveis"])

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 18

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        excel_path = tmp.name
    wb.save(excel_path)

    return FileResponse(excel_path, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        filename=f"relatorio_livros_{date.today().isoformat()}.xlsx")

# ===========================================================================
#  ROTAS — QR CODE
# ===========================================================================

@app.get("/api/qrcode/livro/{livro_id}")
def gerar_qrcode_livro(livro_id: int):
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT id, titulo, autor, isbn FROM livros WHERE id = %s", (livro_id,))
    livro = cursor.fetchone()
    conn.close()
    if not livro:
        raise HTTPException(status_code=404, detail="Livro não encontrado.")
    dados_qr = f"Livro: {livro['titulo']} | Autor: {livro['autor']} | ISBN: {livro['isbn'] or 'N/A'}"
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
    qr.add_data(dados_qr)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="image/png", headers={
        "Content-Disposition": f"inline; filename=qr_livro_{livro_id}.png"
    })

@app.get("/api/qrcode/usuario/{usuario_id}")
def gerar_qrcode_usuario(usuario_id: int):
    conn = obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT id, nome, email, tipo FROM usuarios_biblio WHERE id = %s", (usuario_id,))
    usuario = cursor.fetchone()
    conn.close()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")
    dados_qr = f"Usuário: {usuario['nome']} | E-mail: {usuario['email']} | Tipo: {usuario['tipo']}"
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
    qr.add_data(dados_qr)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="image/png", headers={
        "Content-Disposition": f"inline; filename=qr_usuario_{usuario_id}.png"
    })

# ===========================================================================
#  FUNÇÕES DE E-MAIL E AGENDAMENTO
# ===========================================================================

def enviar_email(destinatario: str, assunto: str, corpo: str):
    try:
        msg = MIMEMultipart()
        msg['From'] = EMAIL_CONFIG["email"]
        msg['To'] = destinatario
        msg['Subject'] = EMAIL_CONFIG["assunto_prefixo"] + assunto
        msg.attach(MIMEText(corpo, 'plain', 'utf-8'))
        server = smtplib.SMTP(EMAIL_CONFIG["smtp_server"], EMAIL_CONFIG["smtp_port"])
        server.starttls()
        server.login(EMAIL_CONFIG["email"], EMAIL_CONFIG["senha"])
        server.sendmail(EMAIL_CONFIG["email"], destinatario, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        logging.error(f"Erro ao enviar e-mail para {destinatario}: {e}")
        return False

def verificar_e_enviar_lembretes():
    print("⏰ Executando verificação de lembretes de devolução...")
    conn = obter_conexao()
    cursor = conn.cursor()
    hoje = date.today()
    data_alerta = hoje + timedelta(days=2)

    cursor.execute("""
        SELECT e.id,
               l.titulo AS titulo_livro,
               u.nome   AS nome_usuario,
               u.email  AS email_usuario,
               e.data_devolucao_prevista,
               DATEDIFF(%s, e.data_devolucao_prevista) AS dias_atraso
          FROM emprestimos e
          JOIN livros l ON l.id = e.id_livro
          JOIN usuarios_biblio u ON u.id = e.id_usuario
         WHERE e.status = 'ativo'
           AND (e.data_devolucao_prevista = %s OR e.data_devolucao_prevista < %s)
    """, (hoje, data_alerta, hoje))

    emprestimos = cursor.fetchall()
    conn.close()

    if not emprestimos:
        print("✅ Nenhum lembrete a ser enviado.")
        return

    for emp in emprestimos:
        if emp["dias_atraso"] > 0:
            status = f"VENCIDO há {emp['dias_atraso']} dias"
        else:
            status = "Vence em 2 dias"

        corpo = f"""
Olá {emp['nome_usuario']},

Este é um lembrete sobre o empréstimo do livro:

📚 Título: {emp['titulo_livro']}
📅 Data de devolução prevista: {emp['data_devolucao_prevista']}
🔔 Status: {status}

Por favor, devolva o livro o mais breve possível para evitar multas.

Atenciosamente,
Equipe AcervoTech
        """
        enviar_email(emp['email_usuario'], f"Lembrete de devolução: {emp['titulo_livro']}", corpo)
        print(f"📧 E-mail enviado para {emp['email_usuario']}")

scheduler = BackgroundScheduler()
scheduler.add_job(
    verificar_e_enviar_lembretes,
    trigger='cron',
    hour=8,
    minute=0,
    id='lembretes_devolucao'
)
scheduler.start()

@app.get("/api/teste/lembretes", summary="Executar manualmente o envio de lembretes")
def executar_lembretes():
    try:
        verificar_e_enviar_lembretes()
        return {"message": "Verificação de lembretes executada com sucesso."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ===========================================================================
#  ARQUIVOS ESTÁTICOS
# ===========================================================================

app.mount("/css", StaticFiles(directory=BASE_DIR / "css"), name="css")
app.mount("/js", StaticFiles(directory=BASE_DIR / "js"), name="js")
app.mount("/pages", StaticFiles(directory=BASE_DIR / "pages"), name="pages")
@app.get("/")
def servir_index():
    return FileResponse(BASE_DIR / "index.html")

# ===========================================================================
#  EXECUÇÃO
# ===========================================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)